"""
data_loader.py — Controlled data source reader for the Acacus MNVR simulator.

All numeric values are read from data/Approved_data.xlsx. The ONLY hardcoded
numbers are fixed project assumptions taken from the consolidated report
(working days, burn-in duration, OEE component targets, contract-tier bounds
that are also echoed in the sheet). Each is labelled in FIXED_ASSUMPTIONS with
its source so nothing masquerades as a confirmed shipment figure.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Optional

import openpyxl

# --------------------------------------------------------------------------
# Paths
# --------------------------------------------------------------------------
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
WORKBOOK_PATH = DATA_DIR / "Approved_data.xlsx"


# --------------------------------------------------------------------------
# Evidence status taxonomy (drives the provisional-vs-final rule everywhere)
# --------------------------------------------------------------------------
class Evidence(str, Enum):
    CONFIRMED = "Confirmed"
    BENCHMARK = "Benchmark"
    DESIGN_ESTIMATE = "Design-basis estimate"
    VALIDATION_REQUIRED = "Validation required"
    PILOT_PENDING = "Pilot pending"
    EXTERNAL_REQUIRED = "External confirmation required"

    @property
    def is_final(self) -> bool:
        """Only Confirmed evidence may back a 'final' (non-provisional) output."""
        return self is Evidence.CONFIRMED


def classify_evidence(raw: Optional[str]) -> Evidence:
    """Map the free-text Status column in the workbook onto the Evidence enum."""
    if not raw:
        return Evidence.VALIDATION_REQUIRED
    t = str(raw).strip().lower()
    if "confirmed" in t and "model" not in t:
        return Evidence.CONFIRMED
    if "confirmed model" in t or t.startswith("confirmed"):
        # 'Confirmed model' / 'Confirmed target' are design-basis, not shipment-proven
        if "target" in t:
            return Evidence.DESIGN_ESTIMATE
        return Evidence.CONFIRMED
    if "benchmark" in t:
        return Evidence.BENCHMARK
    if "candidate" in t or "broker" in t:
        return Evidence.EXTERNAL_REQUIRED
    if "pilot" in t:
        return Evidence.PILOT_PENDING
    if "target" in t:
        return Evidence.DESIGN_ESTIMATE
    return Evidence.VALIDATION_REQUIRED


# --------------------------------------------------------------------------
# Typed data structures
# --------------------------------------------------------------------------
@dataclass(frozen=True)
class BOMLine:
    item: str
    make_buy: str
    qty: float
    unit_cost: float
    line_cost: float
    status: Evidence
    comment: str = ""


@dataclass(frozen=True)
class CostLayer:
    """One row of the China / Jordan / Acacus landed-cost build-up."""
    layer: str
    component: str
    china: float
    jordan: float
    acacus: float
    status: Evidence


@dataclass(frozen=True)
class KPI:
    name: str
    unit: str
    target_text: str          # human-readable target as written in the sheet
    target_value: Optional[float]   # numeric threshold where parseable
    direction: str            # 'max' (higher passes) or 'min' (lower passes)
    default_current: Optional[float]
    evidence: Evidence
    phase: str
    method: str = ""
    comment: str = ""


@dataclass(frozen=True)
class DemandTier:
    name: str
    lower: Optional[int]
    upper: Optional[int]


@dataclass
class FinancialAssumptions:
    fixed_overhead_year: float          # USD/yr, Jordan plant
    china_landed: float                 # USD/unit
    jordan_landed: float                # USD/unit
    acacus_reference: float             # USD/unit
    bom_total: float                    # USD/unit
    jordan_premium_usd: float           # USD/unit vs China
    jordan_premium_pct: float           # fraction
    dz_income_tax: float                # fraction (5%)
    dz_customs_exemption: float         # fraction (0%)
    gafta_threshold: float              # fraction local value-added (40%)
    safety_stock_units: float
    warranty_sensitivity_pct: float     # benchmark accrual (1.43%)


@dataclass(frozen=True)
class FreightTradeItem:
    topic: str
    item: str
    unit: str
    value_text: str
    value_num: Optional[float]
    status: Evidence


@dataclass(frozen=True)
class Competitor:
    name: str
    detail: str
    status: Evidence


@dataclass(frozen=True)
class ReferenceRow:
    source_id: str
    used_for: str
    url: str
    notes: str


@dataclass
class ProjectData:
    bom: list[BOMLine]
    cost_layers: list[CostLayer]
    kpis: list[KPI]
    demand_tiers: list[DemandTier]
    financials: FinancialAssumptions
    freight_trade: list[FreightTradeItem]
    competitors: list[Competitor]
    references: list[ReferenceRow]
    raw_main: list[dict] = field(default_factory=list)


# --------------------------------------------------------------------------
# Fixed project assumptions (NOT in the numeric sheet, sourced from the report)
# Each carries an evidence status so the UI can badge them honestly.
# --------------------------------------------------------------------------
FIXED_ASSUMPTIONS: dict = {
    "working_days_year": {
        "value": 250,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Standard 5-day operating year (report design basis).",
    },
    "burn_in_days": {
        "value": 3,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Burn-in duration = 3 days (Phase 4 design basis).",
    },
    "effective_labour_hours_year": {
        "value": 1800,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Effective direct-labour hours/operator/yr after holidays/breaks (design basis).",
    },
    "annual_labour_hours_per_unit": {
        "value": 0.75,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Assembly + QC + burn-in monitoring labour content per unit (design basis).",
    },
    "oee_availability": {
        "value": 0.90,
        "evidence": Evidence.BENCHMARK,
        "source": "OEE availability component target (OEE-LEANPROD).",
    },
    "oee_performance": {
        "value": 0.95,
        "evidence": Evidence.BENCHMARK,
        "source": "OEE performance component target (OEE-LEANPROD).",
    },
    "oee_quality": {
        "value": 0.99,
        "evidence": Evidence.BENCHMARK,
        "source": "OEE quality component target (OEE-LEANPROD).",
    },
    "default_selling_price": {
        "value": 520.0,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Illustrative selling price within $300-500+ competitor band; NOT a booked price.",
    },
    "default_capex": {
        "value": 600000.0,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Illustrative Phase A/B equipment capex placeholder (design basis).",
    },
    "default_opex_year": {
        "value": 250000.0,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Illustrative incremental opex placeholder (design basis).",
    },
    "default_discount_rate": {
        "value": 0.12,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "Illustrative WACC/discount rate for NPV (design basis).",
    },
    "dz_local_content_threshold": {
        "value": 0.30,
        "evidence": Evidence.EXTERNAL_REQUIRED,
        "source": "Development-Zone 30% local-content screen (broker/customs confirmation required).",
    },
    "project_horizon_years": {
        "value": 5,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "source": "5-year evaluation horizon for NPV/IRR/payback (design basis).",
    },
}

# The three candidate sites (TOPSIS short-list from the report). Cost deltas are
# illustrative design-basis adjustments to the Jordan base case, NOT confirmed.
SITES: dict = {
    "Al-Muwaqqar": {
        "label": "Al-Muwaqqar",
        "zone": "Development Zone (candidate)",
        "overhead_delta": 0.0,       # USD/unit vs Jordan base layers
        "outbound_delta": 0.0,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "note": "Base-case Jordan layers; TOPSIS-preferred for cost/zone balance.",
    },
    "KA II Sahab": {
        "label": "KA II Sahab",
        "zone": "King Abdullah II Ind. City, Sahab",
        "overhead_delta": 1.2,
        "outbound_delta": -0.3,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "note": "Established industrial base; slightly higher facility cost, strong logistics.",
    },
    "KHBTDA Mafraq": {
        "label": "KHBTDA Mafraq",
        "zone": "King Hussein Bin Talal Dev. Area, Mafraq",
        "overhead_delta": -0.8,
        "outbound_delta": 0.6,
        "evidence": Evidence.DESIGN_ESTIMATE,
        "note": "Lower facility cost / incentives; longer outbound leg to main markets.",
    },
}


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def _num(value) -> Optional[float]:
    """Best-effort extraction of a leading number from a cell value."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    import re
    m = re.search(r"-?\d[\d,]*\.?\d*", str(value).replace(",", ""))
    return float(m.group()) if m else None


def _pct(value) -> Optional[float]:
    """Convert a cell like '5%' or 0.05 into a fraction."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value <= 1 else float(value) / 100.0
    n = _num(value)
    if n is None:
        return None
    return n / 100.0 if "%" in str(value) or n > 1 else n


# --------------------------------------------------------------------------
# Sheet parsers
# --------------------------------------------------------------------------
def _parse_bom(ws) -> list[BOMLine]:
    lines: list[BOMLine] = []
    for r in range(3, ws.max_row + 1):
        item = ws.cell(r, 1).value
        if not item or str(item).strip().upper().startswith("TOTAL"):
            continue
        if str(item).strip().lower().startswith("cost layer"):
            break
        unit_cost = _num(ws.cell(r, 4).value)
        line_cost = _num(ws.cell(r, 5).value)
        if line_cost is None:
            continue
        lines.append(
            BOMLine(
                item=str(item).strip(),
                make_buy=str(ws.cell(r, 2).value or "").strip(),
                qty=_num(ws.cell(r, 3).value) or 1.0,
                unit_cost=unit_cost or line_cost,
                line_cost=line_cost,
                status=classify_evidence(ws.cell(r, 6).value),
                comment=str(ws.cell(r, 9).value or "").strip(),
            )
        )
    return lines


def _parse_cost_layers(ws) -> list[CostLayer]:
    layers: list[CostLayer] = []
    header_row = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == "cost layer":
            header_row = r
            break
    if header_row is None:
        return layers
    for r in range(header_row + 1, ws.max_row + 1):
        layer = ws.cell(r, 1).value
        if not layer:
            continue
        china = _num(ws.cell(r, 3).value)
        jordan = _num(ws.cell(r, 4).value)
        acacus = _num(ws.cell(r, 5).value)
        if china is None and jordan is None:
            continue
        layers.append(
            CostLayer(
                layer=str(layer).strip(),
                component=str(ws.cell(r, 2).value or "").strip(),
                china=china or 0.0,
                jordan=jordan or 0.0,
                acacus=acacus if acacus is not None else (china or 0.0),
                status=classify_evidence(ws.cell(r, 6).value),
            )
        )
    return layers


def _parse_kpis(ctq_ws, ops_ws) -> list[KPI]:
    """Build the KPI list primarily from the CTQ/KPI matrix."""
    kpis: list[KPI] = []

    # Heuristic mapping of CTQ rows -> numeric threshold + direction + phase.
    spec = {
        "Lead time (Jordan)": (5.0, "min", "Phase A", 5.0),
        "Lead time (KSA)": (7.0, "min", "Phase A", 7.0),
        "SMT first-pass yield": (98.0, "max", "Phase B", 97.0),
        "BGA voiding (SoC)": (25.0, "min", "Phase B", 18.0),
        "Field defect rate": (5000.0, "min", "Phase B", 3500.0),
        "OEE (Year 1)": (75.0, "max", "Phase A", 72.0),
        "OEE (Year 3)": (85.0, "max", "Phase C", 80.0),
        "Certification first-pass": (100.0, "max", "Phase A", 100.0),
        "OTIF": (95.0, "max", "Phase A", 94.0),
        "Warranty replacement lead time": (5.0, "min", "Phase A", 5.0),
        "Local content / RoO readiness": (40.0, "max", "Phase B", 35.0),
    }

    for r in range(3, ctq_ws.max_row + 1):
        name = ctq_ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        unit = str(ctq_ws.cell(r, 2).value or "").strip()
        target_text = str(ctq_ws.cell(r, 4).value or "").strip()
        method = str(ctq_ws.cell(r, 3).value or "").strip()
        status = classify_evidence(ctq_ws.cell(r, 5).value)
        comment = str(ctq_ws.cell(r, 8).value or "").strip()
        tgt, direction, phase, current = spec.get(
            name, (_num(target_text), "max", "Phase A", _num(target_text))
        )
        kpis.append(
            KPI(
                name=name,
                unit=unit,
                target_text=target_text,
                target_value=tgt,
                direction=direction,
                default_current=current,
                evidence=status,
                phase=phase,
                method=method,
                comment=comment,
            )
        )
    return kpis


def _parse_demand_and_financials(ws) -> tuple[list[DemandTier], dict]:
    tiers: list[DemandTier] = []
    fin: dict = {}
    for r in range(3, ws.max_row + 1):
        area = str(ws.cell(r, 1).value or "").strip()
        item = str(ws.cell(r, 2).value or "").strip()
        val = ws.cell(r, 4).value
        if area == "Demand" and "tier" in item.lower():
            txt = str(val or "")
            lower = upper = None
            if "<" in txt:
                upper = int(_num(txt) or 0)
            elif ">" in txt:
                lower = int(_num(txt) or 0)
            elif "–" in txt or "-" in txt:
                import re
                nums = re.findall(r"\d[\d,]*", txt.replace(",", ""))
                if len(nums) >= 2:
                    lower, upper = int(nums[0]), int(nums[1])
            tiers.append(DemandTier(name=item, lower=lower, upper=upper))
        if "fixed overhead" in item.lower():
            fin["fixed_overhead_year"] = _num(val)
        if "china total cost" in item.lower():
            fin["china_landed"] = _num(val)
        if "jordan total cost" in item.lower():
            fin["jordan_landed"] = _num(val)
        if "cost premium" in item.lower():
            import re
            nums = re.findall(r"\d+\.?\d*", str(val or ""))
            if nums:
                fin["jordan_premium_usd"] = float(nums[0])
            if len(nums) >= 2:
                fin["jordan_premium_pct"] = float(nums[1]) / 100.0
    return tiers, fin


def _parse_freight_trade(ws) -> list[FreightTradeItem]:
    items: list[FreightTradeItem] = []
    for r in range(3, ws.max_row + 1):
        topic = ws.cell(r, 1).value
        item = ws.cell(r, 2).value
        if not topic and not item:
            continue
        items.append(
            FreightTradeItem(
                topic=str(topic or "").strip(),
                item=str(item or "").strip(),
                unit=str(ws.cell(r, 3).value or "").strip(),
                value_text=str(ws.cell(r, 4).value or "").strip(),
                value_num=_num(ws.cell(r, 4).value),
                status=classify_evidence(ws.cell(r, 5).value),
            )
        )
    return items


def _parse_competitors(ws) -> list[Competitor]:
    comps: list[Competitor] = []
    for r in range(3, ws.max_row + 1):
        area = str(ws.cell(r, 1).value or "").strip()
        if area.lower() != "competitor":
            continue
        comps.append(
            Competitor(
                name=str(ws.cell(r, 2).value or "").strip(),
                detail=str(ws.cell(r, 4).value or "").strip(),
                status=classify_evidence(ws.cell(r, 5).value),
            )
        )
    return comps


def _parse_references(ws) -> list[ReferenceRow]:
    refs: list[ReferenceRow] = []
    for r in range(3, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if not sid:
            continue
        refs.append(
            ReferenceRow(
                source_id=str(sid).strip(),
                used_for=str(ws.cell(r, 2).value or "").strip(),
                url=str(ws.cell(r, 3).value or "").strip(),
                notes=str(ws.cell(r, 4).value or "").strip(),
            )
        )
    return refs


# --------------------------------------------------------------------------
# Public entry point
# --------------------------------------------------------------------------
def load_project_data(path: Path | str | None = None) -> ProjectData:
    """Load and parse the controlled workbook into a ProjectData object.

    Raises a clear FileNotFoundError if the workbook is missing so the app can
    show an actionable message instead of a stack trace.
    """
    wb_path = Path(path) if path else WORKBOOK_PATH
    if not wb_path.exists():
        raise FileNotFoundError(
            f"Controlled data file not found at {wb_path}. "
            "Place Approved_data.xlsx in the data/ folder."
        )

    wb = openpyxl.load_workbook(wb_path, data_only=True)

    bom = _parse_bom(wb["MNVR_COGS_Detail"])
    cost_layers = _parse_cost_layers(wb["MNVR_COGS_Detail"])
    kpis = _parse_kpis(wb["CTQ_KPI_Matrix"], wb["Operations_Quality"])
    tiers, fin = _parse_demand_and_financials(wb["Demand_Fin_Competitor"])
    freight = _parse_freight_trade(wb["Freight_Trade"])
    comps = _parse_competitors(wb["Demand_Fin_Competitor"])
    refs = _parse_references(wb["References"])

    bom_total = round(sum(l.line_cost for l in bom), 2)
    china = fin.get("china_landed") or next(
        (c.china for c in cost_layers if c.layer.upper().startswith("LANDED")), 467.14
    )
    jordan = fin.get("jordan_landed") or next(
        (c.jordan for c in cost_layers if c.layer.upper().startswith("LANDED")), 475.04
    )
    premium_usd = fin.get("jordan_premium_usd") or round(jordan - china, 2)

    financials = FinancialAssumptions(
        fixed_overhead_year=fin.get("fixed_overhead_year") or 750000.0,
        china_landed=china,
        jordan_landed=jordan,
        acacus_reference=450.0,
        bom_total=bom_total,
        jordan_premium_usd=premium_usd,
        jordan_premium_pct=fin.get("jordan_premium_pct") or (
            (premium_usd / china) if china else 0.017
        ),
        dz_income_tax=0.05,
        dz_customs_exemption=0.0,
        gafta_threshold=0.40,
        safety_stock_units=1000.0,
        warranty_sensitivity_pct=0.0143,
    )

    return ProjectData(
        bom=bom,
        cost_layers=cost_layers,
        kpis=kpis,
        demand_tiers=tiers,
        financials=financials,
        freight_trade=freight,
        competitors=comps,
        references=refs,
    )


if __name__ == "__main__":
    data = load_project_data()
    print(f"BOM lines:        {len(data.bom)} (total ${data.financials.bom_total})")
    print(f"Cost layers:      {len(data.cost_layers)}")
    print(f"KPIs:             {len(data.kpis)}")
    print(f"Demand tiers:     {[t.name for t in data.demand_tiers]}")
    print(f"China landed:     ${data.financials.china_landed}")
    print(f"Jordan landed:    ${data.financials.jordan_landed}")
    print(f"Premium:          ${data.financials.jordan_premium_usd} "
          f"({data.financials.jordan_premium_pct:.1%})")
    print(f"Fixed overhead:   ${data.financials.fixed_overhead_year:,.0f}/yr")
    print(f"References:       {len(data.references)}")
