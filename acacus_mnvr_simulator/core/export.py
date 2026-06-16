"""
export.py — Scenario export for the Acacus MNVR simulator.

Stage 2 provides a working Excel scenario export. Stage 5 will extend this with
an HTML/PDF executive summary. Everything here returns bytes so the Streamlit
download_button can serve the file without writing to disk.
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.calculations import (
    ScenarioInputs,
    compute_capacity,
    compute_costs,
    compute_finance,
    compute_local_content,
    compute_margin,
)
from core.data_loader import ProjectData

_HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
_LABEL_FONT = Font(bold=True, name="Arial")
_BODY_FONT = Font(name="Arial")
_PROV_FILL = PatternFill("solid", fgColor="FFF2CC")


def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def export_scenario_xlsx(s: ScenarioInputs, data: ProjectData) -> bytes:
    """Build an Excel workbook of the current scenario and return it as bytes."""
    cost = compute_costs(s)
    margin = compute_margin(s, cost)
    cap = compute_capacity(s)
    fin = compute_finance(s, cost, margin)
    lc = compute_local_content(s, cost, data)

    wb = Workbook()

    # ---- Sheet 1: Scenario inputs ----
    ws = wb.active
    ws.title = "Scenario Inputs"
    ws["A1"] = "Acacus MNVR Simulator — Scenario Export"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A2"] = f"Generated: {date.today().isoformat()}"
    ws["A2"].font = _BODY_FONT
    ws.append([])
    ws.append(["Input", "Value"])
    _style_header(ws, 4, 2)
    rows = [
        ("Annual volume (units/yr)", s.annual_volume),
        ("Phase", s.phase),
        ("Site", s.site),
        ("BOM cost (USD/unit)", s.bom_cost),
        ("Scrap / yield loss (%)", f"{s.scrap_pct:.1%}"),
        ("Direct labour (USD/unit)", s.direct_labour),
        ("Mfg overhead (USD/unit)", s.mfg_overhead),
        ("Fixed overhead (USD/yr)", s.fixed_overhead_year),
        ("Inbound freight (USD/unit)", s.inbound_freight),
        ("Outbound freight (USD/unit)", s.outbound_freight),
        ("Sea share (%)", f"{s.sea_share:.0%}"),
        ("Warranty provision (% price)", f"{s.warranty_pct:.2%}"),
        ("Tax rate (%)", f"{s.tax_rate:.1%}"),
        ("Development-Zone benefit", "On" if s.dz_benefit else "Off"),
        ("GAFTA benefit", "On" if s.gafta_benefit else "Off"),
        ("Capex (USD)", s.capex),
        ("Opex (USD/yr)", s.opex_year),
        ("Discount rate (%)", f"{s.discount_rate:.1%}"),
        ("Selling price (USD/unit)", s.selling_price),
        ("China baseline landed (USD/unit)", s.china_baseline_landed),
    ]
    for label, val in rows:
        ws.append([label, val])
        ws.cell(ws.max_row, 1).font = _LABEL_FONT
        ws.cell(ws.max_row, 2).font = _BODY_FONT
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    # ---- Sheet 2: Results ----
    ws2 = wb.create_sheet("Results")
    ws2.append(["Result", "Value", "Note"])
    _style_header(ws2, 1, 3)
    res_rows = [
        ("Absorbed standard COGS (USD/unit)", cost.absorbed_standard_cost, "Standard-cost landed view"),
        ("Fixed overhead per unit (USD/unit)", cost.fixed_overhead_per_unit, f"@ {s.annual_volume:,} units/yr"),
        ("Fully loaded unit cost (USD/unit)", cost.fully_loaded_unit_cost, "Full-overhead-at-volume view"),
        ("Delta vs China — standard (USD/unit)", cost.delta_vs_china_standard, "Absorbed vs China baseline"),
        ("Delta vs China — loaded (USD/unit)", cost.delta_vs_china_loaded, "Fully loaded vs China baseline"),
        ("Total annual cost (USD/yr)", margin.total_annual_cost, "Variable x volume + fixed overhead"),
        ("Unit margin (USD/unit)", margin.unit_margin, "Provisional — illustrative price"),
        ("Annual contribution (USD/yr)", margin.annual_contribution, "Provisional — illustrative price"),
        ("Break-even volume (units/yr)", margin.break_even_volume, margin.break_even_note),
        ("Daily build rate (units/day)", cap.daily_build_rate, f"{cap.working_days} working days"),
        ("Burn-in rack positions", cap.burn_in_rack_positions, "Daily build x 3-day burn-in"),
        ("Burn-in WIP (units)", cap.burn_in_wip, "Daily build x 3-day burn-in"),
        ("Direct operators", cap.direct_operators, "Design-basis labour content"),
        ("Local value added (%)", f"{lc.local_value_added_pct:.1%}", "Provisional — customs ruling required"),
        ("GAFTA 40% threshold", "PASS" if lc.gafta_pass else "FAIL", "Provisional"),
        ("Development-Zone 30% threshold", "PASS" if lc.dz_pass else "FAIL", "Provisional"),
        ("NPV (USD)", fin.npv, "PROVISIONAL — " + fin.note),
        ("IRR", f"{fin.irr:.1%}" if fin.irr is not None else "n/a", "PROVISIONAL"),
        ("Payback (years)", fin.payback_years if fin.payback_years else "n/a", "PROVISIONAL"),
    ]
    for label, val, note in res_rows:
        ws2.append([label, val, note])
        ws2.cell(ws2.max_row, 1).font = _LABEL_FONT
        ws2.cell(ws2.max_row, 2).font = _BODY_FONT
        ws2.cell(ws2.max_row, 3).font = _BODY_FONT
        if "PROVISIONAL" in note or "Provisional" in note:
            for c in range(1, 4):
                ws2.cell(ws2.max_row, c).fill = _PROV_FILL
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 50

    # ---- Sheet 3: Cost build-up ----
    ws3 = wb.create_sheet("Cost Build-up")
    ws3.append(["Cost element", "USD/unit"])
    _style_header(ws3, 1, 2)
    build = [
        ("Material BOM", cost.bom),
        ("Scrap / yield loss", cost.scrap),
        ("Direct labour", cost.direct_labour),
        ("Manufacturing overhead", cost.mfg_overhead),
        ("Inbound freight (blended)", cost.inbound_freight),
        ("Outbound freight", cost.outbound_freight),
        ("Warranty provision", cost.warranty),
        ("= Absorbed standard COGS", cost.absorbed_standard_cost),
        ("+ Fixed overhead / unit", cost.fixed_overhead_per_unit),
        ("= Fully loaded unit cost", cost.fully_loaded_unit_cost),
    ]
    for label, val in build:
        ws3.append([label, round(val, 2)])
        bold = label.startswith("=")
        ws3.cell(ws3.max_row, 1).font = Font(bold=bold, name="Arial")
        ws3.cell(ws3.max_row, 2).font = Font(bold=bold, name="Arial")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ==========================================================================
# HTML executive summary (Stage 5)
# A self-contained HTML document — no external libraries, prints to PDF
# cleanly from any browser (File > Print > Save as PDF). Chosen over a direct
# PDF library for reliability on any laptop with zero extra dependencies.
# ==========================================================================
def export_executive_html(s: ScenarioInputs, data: ProjectData) -> bytes:
    from core.calculations import (
        compute_capacity,
        compute_costs,
        compute_finance,
        compute_local_content,
        compute_margin,
    )
    from core.gates import build_gates, recommend_phase

    cost = compute_costs(s)
    margin = compute_margin(s, cost)
    cap = compute_capacity(s)
    fin = compute_finance(s, cost, margin)
    lc = compute_local_content(s, cost, data)
    gates = build_gates()
    rec = recommend_phase(gates, s.annual_volume)
    f = data.financials

    status_color = {"GO": "#2E8B57", "PARTIAL": "#E8A33D", "HOLD": "#C8553D"}

    def gate_rows():
        out = ""
        for name, g in gates.items():
            col = status_color.get(g.decision.value, "#8A94A6")
            out += (
                f"<tr><td><b>{name}</b> — {g.title}</td>"
                f"<td style='text-align:center'>"
                f"<span style='background:{col};color:#fff;padding:2px 10px;"
                f"border-radius:5px;font-weight:700;font-size:11px'>"
                f"{g.decision.value}</span></td>"
                f"<td>{g.pass_count}/{len(g.checks)} checks</td>"
                f"<td style='font-size:11px;color:#555'>{g.decision_reason}</td></tr>")
        return out

    def cond_items():
        return "".join(f"<li>{c}</li>" for c in rec.headline_conditions)

    rec_col = status_color.get(rec.status.value, "#8A94A6")
    irr_txt = f"{fin.irr:.1%}" if fin.irr is not None else "n/a"
    payback_txt = f"{fin.payback_years} yrs" if fin.payback_years else "beyond horizon"

    html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>Acacus MNVR — Executive Summary</title>
<style>
  @page {{ size: A4; margin: 18mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; color:#1A2233;
         line-height:1.5; max-width:820px; margin:0 auto; padding:24px; }}
  h1 {{ color:#1F3A5F; font-size:24px; margin:0 0 4px 0; }}
  h2 {{ color:#1F3A5F; font-size:16px; margin:22px 0 8px 0;
        border-bottom:2px solid #E6E9EF; padding-bottom:4px; }}
  .sub {{ color:#8A94A6; font-size:13px; margin-bottom:16px; }}
  .prov {{ background:#FFF2CC; border:1px solid #E8A33D; border-radius:8px;
           padding:10px 14px; font-size:12px; color:#7A5A00; margin:14px 0; }}
  .rec {{ background:{rec_col}10; border-left:5px solid {rec_col};
          border-radius:8px; padding:14px 18px; margin:14px 0; }}
  .rec .h {{ font-size:18px; font-weight:800; color:#1A2233; }}
  .cards {{ display:flex; gap:10px; margin:14px 0; flex-wrap:wrap; }}
  .card {{ flex:1; min-width:150px; border:1px solid #E6E9EF; border-radius:8px;
           padding:10px 12px; }}
  .card .l {{ font-size:10px; text-transform:uppercase; color:#5A6478;
              font-weight:700; letter-spacing:0.04em; }}
  .card .v {{ font-size:20px; font-weight:700; color:#1A2233; }}
  .card .s {{ font-size:11px; color:#8A94A6; }}
  table {{ width:100%; border-collapse:collapse; font-size:12px; margin:8px 0; }}
  th {{ background:#1F3A5F; color:#fff; text-align:left; padding:6px 8px; }}
  td {{ border-bottom:1px solid #EEF0F4; padding:6px 8px; vertical-align:top; }}
  ol, ul {{ margin:6px 0 6px 18px; font-size:13px; }}
  .foot {{ margin-top:24px; padding-top:10px; border-top:1px solid #E6E9EF;
           font-size:10px; color:#8A94A6; }}
  .badge {{ background:#FFF2CC; color:#9A6A00; border:1px solid #E8A33D;
            padding:1px 7px; border-radius:7px; font-size:10px; font-weight:700; }}
</style></head><body>

<h1>Acacus MNVR Manufacturing Relocation</h1>
<div class="sub">Feasibility Executive Summary · generated {date.today().isoformat()}
· scenario: {s.site}, {s.annual_volume:,} units/yr, {s.phase}</div>

<div class="rec">
  <div style="font-size:10px;text-transform:uppercase;color:#5A6478;font-weight:700">
  Current recommendation <span class="badge">PROVISIONAL</span></div>
  <div class="h">{rec.headline}</div>
  <div style="font-size:13px;color:#3A4456;margin-top:6px">{rec.rationale}</div>
</div>

<h2>The decision in one view</h2>
<div class="cards">
  <div class="card"><div class="l">China baseline</div>
    <div class="v">${f.china_landed:,.2f}</div><div class="s">Landed, contract mfg</div></div>
  <div class="card"><div class="l">Jordan base case</div>
    <div class="v">${f.jordan_landed:,.2f}</div><div class="s">Standard-cost build-up</div></div>
  <div class="card"><div class="l">High-volume premium</div>
    <div class="v">+${f.jordan_premium_usd:,.2f}</div>
    <div class="s">{f.jordan_premium_pct:.1%} (std cost)</div></div>
  <div class="card"><div class="l">Fully loaded @ {s.annual_volume:,}</div>
    <div class="v">${cost.fully_loaded_unit_cost:,.0f}</div>
    <div class="s">incl. ${cost.fixed_overhead_per_unit:,.0f} fixed OH/unit</div></div>
</div>

<p style="font-size:13px"><b>Cost is close, not cheaper.</b> The +{f.jordan_premium_pct:.1%}
premium is a high-volume standard-cost comparison — not the operative start-up
gap. The relocation case rests on lead time to MENA markets, working-capital
reduction, supply continuity, IP protection and trade origin, not on unit cost.
At low volume the ${f.fixed_overhead_year:,.0f}/yr fixed overhead dominates; the
case improves as volume absorbs it.</p>

<h2>Phase-gate status (evidence-gated, not calendar-gated)</h2>
<table>
  <tr><th>Phase</th><th style="text-align:center">Decision</th>
      <th>Checks</th><th>Reason</th></tr>
  {gate_rows()}
</table>

<h2>{rec.next_phase} proceeds only when these are met together</h2>
<ol>{cond_items()}</ol>

<h2>Operational &amp; capacity snapshot</h2>
<div class="cards">
  <div class="card"><div class="l">Daily build rate</div>
    <div class="v">{cap.daily_build_rate:g}</div><div class="s">units/day</div></div>
  <div class="card"><div class="l">Burn-in WIP</div>
    <div class="v">{cap.burn_in_wip:g}</div><div class="s">units (3-day window)</div></div>
  <div class="card"><div class="l">Direct operators</div>
    <div class="v">{cap.direct_operators:g}</div><div class="s">≈{cap.assembly_stations} stations</div></div>
  <div class="card"><div class="l">Local value added</div>
    <div class="v">{lc.local_value_added_pct:.0%}</div>
    <div class="s">GAFTA 40% / DZ 30%: {'pass' if lc.gafta_pass else 'below'}</div></div>
</div>

<h2>Financial indicators <span class="badge">PROVISIONAL</span></h2>
<div class="cards">
  <div class="card"><div class="l">{s.horizon_years}-yr NPV</div>
    <div class="v">${fin.npv:,.0f}</div><div class="s">@ {s.discount_rate:.0%}</div></div>
  <div class="card"><div class="l">IRR</div>
    <div class="v">{irr_txt}</div><div class="s">provisional</div></div>
  <div class="card"><div class="l">Payback</div>
    <div class="v">{payback_txt}</div><div class="s">undiscounted</div></div>
  <div class="card"><div class="l">Assumed price</div>
    <div class="v">${s.selling_price:,.0f}</div><div class="s">illustrative</div></div>
</div>
<div class="prov"><b>⚠ Read before quoting these figures.</b> Selling price,
capex, opex and discount rate are design-basis placeholders, not booked figures.
NPV, IRR and payback move sharply with the assumed selling price and must not be
presented as final. All financial and trade benefits are provisional pending
evidence and a customs ruling.</div>

<div class="foot">
  Data source: Approved_data.xlsx (controlled). This summary reflects one
  scenario and the current gate state; figures recompute live in the simulator.
  Phases are evidence-gated — each phase is a gate, not a date.
</div>
</body></html>"""
    return html.encode("utf-8")


if __name__ == "__main__":
    from core.calculations import default_inputs
    from core.data_loader import load_project_data
    d = load_project_data()
    s = default_inputs(d)
    blob = export_scenario_xlsx(s, d)
    with open("/tmp/scenario_test.xlsx", "wb") as f:
        f.write(blob)
    print(f"Wrote {len(blob):,} bytes to /tmp/scenario_test.xlsx")
